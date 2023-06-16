from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from typing import Optional, Union
import enum

class DischargeVar(enum.Enum):
    DOD = enum.auto()
    SOC = enum.auto()
    MAH = enum.auto()
    AH = enum.auto()

class DischargeData():
    
    def __init__(self, file: Union[Path, str], x_axis: DischargeVar, nominal_capacity_Ah: float,
                 dod_lower: float=0, dod_upper: float=1):
        """initialize the discharge data

        Parameters
        ----------
        file : Union[Path, str]
            File path to excel spreadsheet containing discharge data. Excel file should
            contain multiple sheets with name format of the form "c_rate 0.5". Each sheet should 
            contain two columns -- The first column must contain the amount the cell has been discharged.
            This can be given as a discharge capacity in mAh or Ah, or directly in terms of SoC or DoD. 
            The second column is the cells terminal voltage. Column names don't matter since the type of
            data in the first column is specificied directly by the 'x_axis' argument
            
        nominal_capacity_Ah: float
            Nominal cell capacity in Ah, used to scale C-rates to current values and 
            discharge capacity to DoD
            
        x_axis : DischargeVar
            A DischargeVar enum specifying the type of data contained in the first column of the excel
            spreadsheet. Valid values are: DischargeVar.SOC, DischargeVar.DOD, DischargeVar.AH, and DischargeVar.MAH
            
        dod_lower: float
            lower limit for cropping depth of discharge data for model extraction
            
        dod_upper: float
            upper limit for cropping depth of discharge data for model extraction
        """
        
        #specify data types for attributes set in "load_data" method
        self._data: pd.DataFrame
        self.nominal_capacity_Ah: float
        
        #load data from excel file
        self.load_data(file=file, nominal_capacity_Ah=nominal_capacity_Ah, discharge_var=x_axis) #load data into _data attribute

        #store dod limits for cropped fit data
        self.dod_lower = dod_lower
        self.dod_upper = dod_upper
        
    def load_data(self, file: Union[Path, str], nominal_capacity_Ah: float, discharge_var: DischargeVar) -> None:
        """load excel file containing multiple sheets into single pandas DataFrame of long format

        Parameters
        ----------
        file : Path or str
            Path to Excel file containing discharge data

        Returns
        -------
        pd.DataFrame
            DataFrame containing all of the discharge curves in long format
        """
        
        #ensure that it is a Path object
        file = Path(file)
        self.nominal_capacity_Ah = nominal_capacity_Ah
        
        #first load the workbook and extract the sheetnames
        wb = load_workbook(filename=file)
        sheets = wb.sheetnames
        wb.close()
        
        #create empty list to dataframes for each excel sheet to
        dfs = []
        
        #loop over each sheet in the excel file
        for sheet in sheets:
            #extract c-rate from sheet name (this assumes sheets are named with convention "c_rate 1.2" or similar)
            c_rate = float(sheet.split()[-1])
            #load data into pandas dataframe (we are overwriting the column names specified in the spreadsheet)
            #the spreadsheet must have a first row of capacity in Ah or DoD and the second column of voltage
            
            #load data with generic x-column name (this will be renamed)
            df = pd.read_excel(file, sheet_name=sheet, names=['x_col', 'V'])
            
            if discharge_var is DischargeVar.SOC:
                df = df.rename(columns={'x_col':'SoC'})
                df['DoD'] = 1 - df['SoC']
                
            elif discharge_var is DischargeVar.DOD:
                df = df.rename(columns={'x_col':'DoD'})
                df['SoC'] = 1 - df['DoD']
                
            elif discharge_var is DischargeVar.MAH:
                df['DoD'] = df['x_col']/(self.nominal_capacity_Ah*1000)
                df['SoC'] = 1 - df['DoD']
                df = df.drop(columns='x_col')
                
            elif discharge_var is DischargeVar.AH:
                df['DoD'] = df['x_col']/(self.nominal_capacity_Ah)
                df['SoC'] = 1 - df['DoD']
                df = df.drop(columns='x_col')
                
            else:
                raise ValueError('Invalid value for discharge_var')
            
            #add additional computed columns to the dataframe
            df['C-rate'] = c_rate
            df['I [A]'] = df['C-rate']*self.nominal_capacity_Ah
            
            dfs.append(df)
            
        #combine all of the dataframes into a single dataframe
        self._data = pd.concat(dfs, ignore_index=True)
    
    @property
    def data(self) -> pd.DataFrame:
        return self._data

    @property
    def data_cropped(self) -> pd.DataFrame:
        """extract subset of all data in limited range of DoD

        Parameters
        ----------
        dod_lower : float
            lower bound of DoD to chop data for fitting
        dod_upper : float
            upper bound of DoD to chop data for fitting
        """
        return self.data[(self.data['DoD']>self.dod_lower) & (self.data['DoD']<self.dod_upper)].copy()
    
    
    def plot(self, cropped=False, **kwargs) -> tuple:
        """convenience function to plot the raw constant current discharge curves

        Parameters
        ----------
        kwargs : dict
            key word arguments to supply to the subplots command in matplotlib

        Returns
        -------
        tuple
            tuple containing the figure and axes objects created
        """
        
        if cropped:
            data = self.data_cropped
        else:
            data = self.data
        
        fig, ax = plt.subplots(**kwargs)
        for (c_rate, df) in data.groupby('C-rate'):
            ax.plot(df['DoD'], df['V'], label=f'{c_rate} C')
        
        ax.set_xlabel('DoD [-]')
        ax.set_ylabel('Terminal Voltage [V]')
        ax.legend()
        
        fig.tight_layout()
    
        return fig, ax
    
if __name__ == '__main__':
    d = DischargeData('inputs/cell_data.xlsx', nominal_capacity_Ah=2.2, x_axis=DischargeVar.DOD)
    d.plot(cropped=True)